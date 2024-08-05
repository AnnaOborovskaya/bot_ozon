import requests
import os
from dotenv import load_dotenv
from typing import Text
import requests
import json
import datetime,time
from datetime import timedelta, datetime
import openpyxl

load_dotenv()

def func_b():
    print(1)
    try:
        url = 'https://api-seller.ozon.ru/v1/actions' #Список акций
        headers = {'Client-Id':os.getenv('client_id'), 'Api-Key':os.getenv('client_secret'),}

        r = requests.get(url, headers = headers)


        keys = {'id':'Индентификатор акции', 'title':'Название акции', 'action_type':'Тип акции', 'description':'Описание акции', 'date_start':'Дата начала акции', 'date_end':'Дата окончания акции',
                'freeze_date':'Дата приостановки акции', 'potential_products_count':'Количество товаров, доступных для акции', 'participating_products_count':'Количество товаров, которые участвуют в акции', 
                'is_participating':'Участвуете вы в этой акции или нет', 'is_voucher_action':'Признак, что для участия в акции покупателям нужен промокод', 'banned_products_count':'Количество заблокированных товаров',
                'with_targeting':'Признак, что акция с целевой аудиторией', 'order_amount':'Сумма заказа', 'discount_type':'Тип скидки', 'discount_value':'Размер скидки'}

        file = open('Promos.txt', 'w')
        #********************список акций
        i = 1
        l = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        path = "C:/Users/Evgen/YandexDisk-promebelmarket/Маркетплейсы/API Аналитика/Озон/Общая модель.xlsx"
        wb = openpyxl.load_workbook(path)  
        sheet = wb['список']
        sheet.delete_rows(2, amount=100)

        important_ar = {}

        for txt in r.json()['result']:
            important_ar[txt["id"]] = txt["title"]
            file.write(f'{txt["id"]}|')
            c = 0
            if i == 1:        
                for key, item in txt.items():
                    sheet[f'{l[c]}{i}'] = keys[key]
                    c += 1
                i += 1                
            else:
                for key, item in txt.items():
                    try:
                        new_date = datetime.strptime(item[:10], '%Y-%m-%d')
                        sheet[f'{l[c]}{i}'] = datetime.strftime(new_date, '%d.%m.%Y')
                    except:                
                        if item == True and l[c] == 'F':
                            sheet[f'{l[c]}{i}'] = 'участвует'
                        elif item == False and l[c] == 'F':
                            sheet[f'{l[c]}{i}'] = 'нет'
                        else:
                            sheet[f'{l[c]}{i}'] = item    
                    c += 1
                i += 1
        file.close()
        wb.save("C:/Users/Evgen/YandexDisk-promebelmarket/Маркетплейсы/API Аналитика/Озон/Общая модель.xlsx")

        #********************список участвующих
        url = 'https://api-seller.ozon.ru/v1/actions/products'  #Список участвующих в акции товаров
        headers = {'Client-Id':os.getenv('client_id'), 'Api-Key':os.getenv('client_secret'),}

        lis = ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

        off = requests.post('https://api-seller.ozon.ru/v4/product/info/prices', headers = headers, json={
                "filter": {
                    "offer_id": [],
                    "product_id": [],
                    "visibility": "ALL"
                },
                "last_id": "",
                "limit": 1000
            })
        offer_id_ar = {}
        for row in off.json()['result']['items']:
            offer_id_ar[row['product_id']] = row['offer_id']  

        etalon_prices = {}
        for row in off.json()['result']['items']:
            etalon_prices[row['product_id']] = row['price']['min_price']


        for_last = {}
        for_last_2 = {}
        wb = openpyxl.load_workbook(path)  
        sheet = wb['Участвуют и доступные для участ']
        i = len([row for row in sheet.values if row[0]])
        f = open('Promos.txt','r')
        for line in f.readline().split('|')[:-1]:
            l = int(line)  
            r = requests.post(url, headers = headers, json={"action_id": l, "limit": 1000, "offset": 0})      
            s = set()
            file = open('prod.txt', 'a')          

            for txt in r.json()["result"]["products"]:                  
                try:
                    i += 1
                    if txt["id"] not in s:
                        file.write(f'{txt["id"]}|')  
                        s.add(id)     
                    c = 0  
                    for key, item in txt.items():   
                        sheet[f'{lis[c]}{i}'] = item
                        c += 1  

                    sheet[f'A{i}'] = datetime.now().strftime('%d.%m.%Y')
                    sheet[f'B{i}'] = 'Участвует'
                    sheet[f'C{i}'] = line
                    sheet[f'D{i}'] = important_ar[int(line)]
                    for_last[txt["id"]] = ['Участвует', important_ar[int(line)]]

                    if txt["id"] in for_last_2: 
                        if ['Участвует', important_ar[int(line)]] not in for_last_2[txt["id"]]:
                            for_last_2[txt["id"]].append(['Участвует', important_ar[int(line)]])
                    else:
                        for_last_2[txt["id"]] = [['Участвует', important_ar[int(line)]]]
                    sheet[f'E{i}'] = offer_id_ar[txt["id"]] 


                    etalon = float(etalon_prices[txt["id"]]) 
                    sheet[f'N{i}'] = float(txt["max_action_price"]) - float(etalon_prices[txt["id"]]) 
                    sheet[f'M{i}'] = float(etalon)       
                except:
                    pass            

        file.close()
        wb.save("C:/Users/Evgen/YandexDisk-promebelmarket/Маркетплейсы/API Аналитика/Озон/Общая модель.xlsx")      

        #********************список доступных
        path = "C:/Users/Evgen/YandexDisk-promebelmarket/Маркетплейсы/API Аналитика/Озон/Общая модель.xlsx"
        wb = openpyxl.load_workbook(path)  
        sheet = wb['Участвуют и доступные для участ']
        i = len([row for row in sheet.values if row[0]])
        f = open('Promos.txt','r')
        for line in f.readline().split('|')[:-1]:
            try:
                url = 'https://api-seller.ozon.ru/v1/actions/candidates' #Список доступных для акции товаров
                l = int(line)  
                r = requests.post(url, headers = headers, json={"action_id": l, "limit": 1000, "offset": 0})     
                s = set()
                file = open('cand.txt', 'a')

                for txt in r.json()["result"]["products"]:
                    try:
                        i += 1 
                        if txt["id"] not in s:
                            file.write(f'{txt["id"]}|')  
                            s.add(id)     
                        c = 0  
                        for key, item in txt.items():
                            sheet[f'{lis[c]}{i}'] = item
                            c += 1
                        sheet[f'A{i}'] = datetime.now().strftime('%d.%m.%Y')
                        sheet[f'B{i}'] = 'Доступен'
                        sheet[f'C{i}'] = line
                        sheet[f'D{i}'] = important_ar[int(line)]
                        for_last[txt["id"]] = ['Доступен', important_ar[int(line)]]
                        sheet[f'E{i}'] = offer_id_ar[txt["id"]]  


                        if txt["id"] in for_last_2: 
                            if ['Доступен', important_ar[int(line)]] not in for_last_2[txt["id"]]:
                                for_last_2[txt["id"]].append(['Доступен', important_ar[int(line)]])
                        else:
                            for_last_2[txt["id"]] = [['Доступен', important_ar[int(line)]]]
                        etalon = float(etalon_prices[txt["id"]]) 
                        sheet[f'N{i}'] = float(txt["max_action_price"]) - float(etalon_prices[txt["id"]]) 
                        sheet[f'M{i}'] = float(etalon)     
                    except:
                        pass            
            except:
                pass
        file.close()
        wb.save("C:/Users/Evgen/YandexDisk-promebelmarket/Маркетплейсы/API Аналитика/Озон/Общая модель.xlsx")

        #**************инфа о ценах
        url = 'https://api-seller.ozon.ru/v4/product/info/prices' #Получить информацию о цене товара
        headers = {'Client-Id':os.getenv('client_id'), 'Api-Key':os.getenv('client_secret'),}


        wb = openpyxl.load_workbook(path)  
        sheet2 = wb['Содержательная бизнес логика']

        i = len([row for row in sheet2.values if row[1]])

        r = requests.post(url, headers = headers, json={
                "filter": {
                    "offer_id": [],
                    "product_id": [],
                    "visibility": "ALL"
                },
                "last_id": "",
                "limit": 1000
            })

        path_new = "C:/Users/Evgen/YandexDisk-promebelmarket/Маркетплейсы/API Аналитика/API ЗАКАЗЫ основной/2024 год/Справочник номенклатур.xlsx" #
        wb_new = openpyxl.load_workbook(path_new)  
        sheet_new = wb_new['Справочник номенклатуры']
        sheet_new_d = {}
        sheet_new_e = {}

        for row in sheet_new.values:
            sheet_new_d[row[1]] = row[5]
            sheet_new_e[row[1]] = f"{row[5]} {row[6]} {row[7]}"

        for row in r.json()['result']['items']:
            
                for text_act in for_last_2[row['product_id']]:
                    try:
                        i += 1
                        sheet2[f'A{i}'] = datetime.now().strftime('%d.%m.%Y')
                        try:
                            sheet2[f'B{i}'] = row['product_id']
                        except:
                            pass    
                        try:
                            sheet2[f'C{i}'] = row['offer_id']
                        except:
                            pass   
                        try:   
                            sheet2[f'D{i}'] = sheet_new_d[row['offer_id']]
                        except:
                            pass 
                        try:
                            sheet2[f'E{i}'] = sheet_new_e[row['offer_id']]
                        except:
                            pass
                        try:    
                            sheet2[f'F{i}'] = row['commissions']['sales_percent_fbo']
                        except:
                            pass 
                        try:   
                            sheet2[f'G{i}'] = row['commissions']['sales_percent_fbs']
                        except:
                            pass
                        try:
                            if row['price']['auto_action_enabled'] == True:
                                sheet2[f'H{i}'] = 'подключены'
                            else:
                                sheet2[f'H{i}'] = 'нет' 
                        except:
                            pass
                        try:                      
                            sheet2[f'I{i}'] = float(row['price']['old_price'])
                        except:
                            pass  
                    

                        try:                      
                            sheet2[f'J{i}'] = float(row['price']['price'])
                        except:
                            pass 
                        try:                      
                            sheet2[f'K{i}'] = float(row['price']['marketing_seller_price'])
                        except:
                            pass 
                        try:                      
                            sheet2[f'L{i}'] = float(row['price']['marketing_price'])
                        except:
                            pass 

                        try:                      
                            sheet2[f'M{i}'] = float(row['price']['min_price'])
                        except:
                            pass 
                        try:                      
                            sheet2[f'N{i}'] = float(row['price_indexes']['price_index'])
                        except:
                            pass 
                        try:                      
                            sheet2[f'O{i}'] = float(row['price_indexes']['external_index_data']['minimal_price'])
                        except:
                            pass 
                        try:                      
                            sheet2[f'P{i}'] = float(row['price_indexes']['external_index_data']['price_index_value'])
                        except:
                            pass 
                        try:                      
                            sheet2[f'Q{i}'] = float(row['price_indexes']['ozon_index_data']['minimal_price'])
                        except:
                            pass 

                        try:                      
                            sheet2[f'R{i}'] = float(row['price_indexes']['ozon_index_data']['price_index_value'])
                        except:
                            pass 
                        try:                      
                            sheet2[f'S{i}'] = float(row['price_indexes']['self_marketplaces_index_data']['minimal_price'])
                        except:
                            pass 
                        try:                      
                            sheet2[f'T{i}'] = text_act[1]
                            sheet2[f'U{i}'] = text_act[0]
                        except:
                            pass    
                    

                    except:
                        pass 

        wb.save("C:/Users/Evgen/YandexDisk-promebelmarket/Маркетплейсы/API Аналитика/Озон/Общая модель.xlsx")  

    except:
        file = open('Ошибки.txt', 'a')
        file.write(f"{datetime.now().strftime('%d.%m.%Y %HS%M:%S')} Ошибка, вероятно нет подключения к интернету\n")